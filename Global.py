import xlsxwriter
import os.path
import logging
import re
from os import path
import openpyxl as xl;
class Global ():
    def __init__(self, answerLength, answers):
        self.questionsInPoll = [] #1 poll için soru texti, kimlerin hangi şıkkı kaç kez seçtiği, seçilen tüm şıklar
        self.questionText = []
        self.answers = answers # soruların cevap anahtarı- doğru cevapları
        self.answerLength = answerLength


    def writeinGlobal(self,students,readPath,Polls,globalFileName,globalFileTitles,totalpoint,dateofpoll):

       if (path.exists(globalFileName)):
           totalquestions = 0
           filename = globalFileName
           wb1 = xl.load_workbook(filename)
           ws1 = wb1.worksheets[0]

           # opening the destination excel file

           filename1 = globalFileName
           wb2 = xl.load_workbook(filename1)
           ws2 = wb2.active

           colunm = ws1.cell(1,200).value

           y=1;
           t=1;

           # add_sheet is used to create sheet.

           while y<=len(students):
               while t<colunm:
                   c=ws1.cell(y,t).value
                   ws2.cell(y,t).value=c
                   t=t+1
               y=y+1

           y=2
           l=0
           while y<=len(students)+1:
               c=ws1.cell(y,colunm).value;
               c=c.split(";")
               totalquestions=int(c[0])
               totalpoint[l]=int(c[1])
               l+=1
               y=y+1

           logging.info("Global file for is created")


           i = 0;
           k=2;
           iter=0;
           length_Polls = len(Polls)
           while iter < length_Polls:
                totalquestions += Polls[iter].answerLength
                newPath = readPath.split(".csv")
                if iter == 0:
                    picname = newPath[0]
                pollName = re.sub("[^0-9a-zA-Z]+", "_", Polls[iter].pollName)
                dateofpoll1 = re.sub("[^0-9a-zA-Z]+", "_", dateofpoll[iter])
                readPath1= "Poll_" + str(Polls[iter].pollNumber) + "_" + pollName + "_" + dateofpoll1
                ws2.cell(1, colunm).value=readPath1
                while i < len(students):
                   totalpoint[i] += students[i].totalpointforThispoll[iter]
                   ws2.cell(k,colunm).value=(str(Polls[iter].answerLength)+" ; "+str((students[i].totalpointforThispoll[iter] / Polls[iter].answerLength) * 100))
                   i=i+1
                   k=k+1
                i=0
                k=2
                colunm+=1
                iter += 1

           e = 0
           k = 2
           ws2.cell(1, colunm).value = 'TOTAL'
           while e < len(students):
               ws2.cell(k, colunm).value=(str(totalquestions) + " ; " + str(totalpoint[e]) +" ; "+str(round((totalpoint[e]*100/totalquestions),2)))
               e += 1
               k += 1
           colunm += 1
           ws2.cell(1,200).value=colunm-1
           wb2.save(globalFileName)
       else:
          colunm=3
          wb = xlsxwriter.Workbook(globalFileName)
          sheet1 = wb.add_worksheet()

        # add_sheet is used to create sheet.

          sheet1.write(0, 0, globalFileTitles[0])
          sheet1.write(0, 1, globalFileTitles[1])
          sheet1.write(0, 2, globalFileTitles[2])
          d = 0
          i = 1
          while d < len(students):
                sheet1.write(i, 0, str(students[d].studentid))
                sheet1.write(i, 1, students[d].firstName)
                sheet1.write(i, 2, students[d].lastName)
                d = d + 1
                i = i + 1

          i = 0;
          k=1
          iter = 0
          totalquestions=0
          length_Polls = len(Polls)
          while iter < length_Polls:
            totalquestions+=Polls[iter].answerLength
            newPath = readPath.split(".csv")
            if iter == 0:
                picname = newPath[0]
            pollName = re.sub("[^0-9a-zA-Z]+", "_", Polls[iter].pollName)
            dateofpoll1 = re.sub("[^0-9a-zA-Z]+", "_", dateofpoll[iter])
            readPath1 = "Poll_" + str(Polls[iter].pollNumber) + "_" + pollName + "_" + dateofpoll1
            sheet1.write(0, colunm, readPath1)
            while i < len(students):
                   totalpoint[i]+=students[i].totalpointforThispoll[iter]
                   sheet1.write(k, colunm, str(Polls[iter].answerLength) + " ; " + str((students[i].totalpointforThispoll[iter] / Polls[iter].answerLength) * 100))
                   i = i + 1
                   k=k+1
            k=1
            i=0
            colunm += 1
            iter += 1

          e=0
          k=1
          sheet1.write(0, colunm, "TOTAL")
          while e < len(students):
              sheet1.write(k, colunm,str(totalquestions)+" ; "+str(totalpoint[e])+"  ;  "+str(round((totalpoint[e]*100/totalquestions),2)))
              e+=1
              k+=1
          colunm+=1
          sheet1.write(0, 199, colunm)
          wb.close()

    def studentınformation(self,Polls,students,readStudentTitles):
        p=0
        directory="./STUDENTS/"
        if not os.path.exists(directory):
            os.makedirs(directory)
        while p< len(students):
           filename=directory+students[p].studentid+" "+students[p].firstName+" "+students[p].lastName+".xlsx"

           if (path.exists(filename)):
               wb2 = xl.load_workbook(filename)
               i = 0

               while i < len(Polls):
                  pollName = re.sub("[^0-9a-zA-Z]+", "_", Polls[i].pollName)
                  readPath = "Poll_" + str(Polls[i].pollNumber) + "_" + pollName
                  sheet1 = wb2.create_sheet(readPath)
                  k = 0
                  l = 2
                  sheet1.cell(1, 1).value= readStudentTitles[0]
                  sheet1.cell(1, 2 ).value=readStudentTitles[1]
                  sheet1.cell(1, 3 ).value=readStudentTitles[2]
                  while k < len(Polls[i].questionText):
                     sheet1.cell(l, 1).value= Polls[i].questionText[k]
                     sheet1.cell(l, 2).value= str(Polls[i].answers[k].answer)
                     if students[p].answerof[i]:
                        sheet1.cell(l, 3).value= students[p].answerof[i][k]
                     sheet1.cell(l, 4).value = students[p].question[i][k]
                     k = k + 1
                     l = l + 1

                  i = i + 1
               wb2.save(filename)
               wb2.close()


           else:
               wb = xlsxwriter.Workbook(filename)
               i=0

               while i<len(Polls):
                 pollName = re.sub("[^0-9a-zA-Z]+", "_", Polls[i].pollName)
                 readPath = "Poll_" + str(Polls[i].pollNumber) + "_" + pollName

                 if len(readPath) > 31:
                     readPath=readPath[:31]
                 sheet1 = wb.add_worksheet(readPath)
                 k=0
                 l=1
                 sheet1.write(0, 0, readStudentTitles[0])
                 sheet1.write(0, 1, readStudentTitles[1])
                 sheet1.write(0, 2, readStudentTitles[2])

                 while k <len(Polls[i].questionText):
                    sheet1.write(l, 0,Polls[i].questionText[k])
                    sheet1.write(l,1,str(Polls[i].answers[k].answer))
                    if  students[p].answerof[i]:
                       sheet1.write(l, 2,students[p].answerof[i][k])
                    sheet1.write(l,3,students[p].question[i][k])
                    k=k+1
                    l=l+1

                 i=i+1

               wb.close()
           p=p+1